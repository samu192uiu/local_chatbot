# services/reminders.py
"""
Sistema de lembretes automáticos para agendamentos.
Envia notificações 1 dia antes e 2 horas antes do horário agendado.
"""

import os
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

from services import excel_services as es

logger = logging.getLogger(__name__)

# Flag para controlar se o scheduler está rodando
_scheduler = None
_waha_send_function = None


def inicializar_lembretes(send_function):
    """
    Inicializa o sistema de lembretes automáticos.
    
    Args:
        send_function: Função para enviar mensagens via WAHA
                      Assinatura: send_function(chat_id: str, mensagem: str)
    """
    global _scheduler, _waha_send_function
    
    _waha_send_function = send_function
    
    if _scheduler is not None:
        logger.warning("Scheduler de lembretes já está rodando")
        return
    
    _scheduler = BackgroundScheduler()
    
    # Job 1: Lembretes de 1 dia antes - rodar às 18h todos os dias
    _scheduler.add_job(
        func=enviar_lembretes_1_dia,
        trigger=CronTrigger(hour=18, minute=0),
        id='lembretes_1_dia',
        name='Lembretes 1 dia antes',
        replace_existing=True
    )
    
    # Job 2: Lembretes de 2h antes - rodar a cada hora
    _scheduler.add_job(
        func=enviar_lembretes_2_horas,
        trigger=CronTrigger(minute=0),  # A cada hora cheia
        id='lembretes_2_horas',
        name='Lembretes 2 horas antes',
        replace_existing=True
    )
    
    _scheduler.start()
    logger.info("✅ Sistema de lembretes iniciado com sucesso")
    logger.info("📅 Lembretes 1 dia antes: Diariamente às 18h")
    logger.info("⏰ Lembretes 2 horas antes: A cada hora")


def parar_lembretes():
    """Para o sistema de lembretes."""
    global _scheduler
    
    if _scheduler:
        _scheduler.shutdown()
        _scheduler = None
        logger.info("🛑 Sistema de lembretes parado")


def enviar_lembretes_1_dia():
    """
    Envia lembretes para agendamentos que ocorrerão amanhã.
    Executado diariamente às 18h.
    """
    try:
        logger.info("📅 Verificando agendamentos para amanhã...")
        
        amanha = datetime.now() + timedelta(days=1)
        data_amanha = amanha.strftime("%d/%m/%Y")
        
        agendamentos = _buscar_agendamentos_por_data(data_amanha)
        
        enviados = 0
        for ag in agendamentos:
            if _deve_enviar_lembrete(ag, tipo="1_dia"):
                sucesso = _enviar_lembrete_1_dia(ag)
                if sucesso:
                    enviados += 1
        
        logger.info(f"✅ Lembretes 1 dia: {enviados} enviados para {len(agendamentos)} agendamentos")
    
    except Exception as e:
        logger.error(f"❌ Erro ao enviar lembretes de 1 dia: {e}")


def enviar_lembretes_2_horas():
    """
    Envia lembretes para agendamentos que ocorrerão em 2 horas.
    Executado a cada hora.
    """
    try:
        logger.info("⏰ Verificando agendamentos nas próximas 2 horas...")
        
        agora = datetime.now()
        daqui_2h = agora + timedelta(hours=2)
        
        # Buscar agendamentos de hoje e amanhã
        data_hoje = agora.strftime("%d/%m/%Y")
        data_amanha = (agora + timedelta(days=1)).strftime("%d/%m/%Y")
        
        agendamentos_hoje = _buscar_agendamentos_por_data(data_hoje)
        agendamentos_amanha = _buscar_agendamentos_por_data(data_amanha)
        
        todos_agendamentos = agendamentos_hoje + agendamentos_amanha
        
        enviados = 0
        for ag in todos_agendamentos:
            # Verificar se está na janela de 2h
            if _esta_em_2_horas(ag, agora, daqui_2h):
                if _deve_enviar_lembrete(ag, tipo="2_horas"):
                    sucesso = _enviar_lembrete_2_horas(ag)
                    if sucesso:
                        enviados += 1
        
        if enviados > 0:
            logger.info(f"✅ Lembretes 2h: {enviados} enviados")
    
    except Exception as e:
        logger.error(f"❌ Erro ao enviar lembretes de 2 horas: {e}")


def _buscar_agendamentos_por_data(data_str: str) -> List[Dict[str, Any]]:
    """Busca agendamentos confirmados de uma data específica."""
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(es.FILE_PATH)
        ws = wb[es.SHEET_AG]
        
        es._ensure_headers(ws, es.HEADERS_AG)
        hm = es._get_header_map(ws)
        
        c_data = hm.get("Data")
        c_status = hm.get("Status")
        
        if not all([c_data, c_status]):
            return []
        
        agendamentos = []
        for r in range(2, ws.max_row + 1):
            data_row = str(ws.cell(row=r, column=c_data).value or "").strip()
            status_row = str(ws.cell(row=r, column=c_status).value or "").strip().lower()
            
            # Apenas agendamentos confirmados da data especificada
            if data_row == data_str and status_row == "confirmado":
                ag = es._row_to_dict(ws, r)
                ag['_row'] = r  # Salvar número da linha para atualização
                agendamentos.append(ag)
        
        return agendamentos
    
    except Exception as e:
        logger.error(f"Erro ao buscar agendamentos de {data_str}: {e}")
        return []


def _esta_em_2_horas(ag: Dict[str, Any], agora: datetime, daqui_2h: datetime) -> bool:
    """Verifica se o agendamento está na janela de 2 horas."""
    try:
        data_str = ag.get("Data", "")
        hora_str = ag.get("Hora", "")
        
        data_hora = datetime.strptime(f"{data_str} {hora_str}", "%d/%m/%Y %H:%M")
        
        # Está entre agora+1h50min e agora+2h10min (janela de 20min)
        inicio_janela = agora + timedelta(hours=1, minutes=50)
        fim_janela = agora + timedelta(hours=2, minutes=10)
        
        return inicio_janela <= data_hora <= fim_janela
    
    except Exception:
        return False


def _deve_enviar_lembrete(ag: Dict[str, Any], tipo: str) -> bool:
    """
    Verifica se deve enviar lembrete (não enviar duplicados).
    
    Args:
        ag: Dicionário do agendamento
        tipo: "1_dia" ou "2_horas"
    """
    lembrete_enviado = ag.get("LembreteEnviado", "")
    
    # Se nunca enviou, pode enviar
    if not lembrete_enviado:
        return True
    
    try:
        # Verificar quando foi o último lembrete
        ultimo_lembrete = datetime.strptime(lembrete_enviado, "%Y-%m-%d %H:%M:%S")
        
        if tipo == "1_dia":
            # Lembrete de 1 dia: só enviar se não enviou nas últimas 20h
            diferenca = datetime.now() - ultimo_lembrete
            return diferenca.total_seconds() > (20 * 3600)
        
        elif tipo == "2_horas":
            # Lembrete de 2h: só enviar se não enviou na última 1h
            diferenca = datetime.now() - ultimo_lembrete
            return diferenca.total_seconds() > 3600
        
    except Exception:
        return True
    
    return True


def _enviar_lembrete_1_dia(ag: Dict[str, Any]) -> bool:
    """Envia lembrete de 1 dia antes."""
    try:
        chat_id = ag.get("ChatId", "")
        nome = ag.get("ClienteNome", "Cliente")
        data = ag.get("Data", "")
        hora = ag.get("Hora", "")
        
        if not chat_id:
            return False
        
        # Formatar nome (só primeiro nome)
        primeiro_nome = nome.split()[0] if nome else "Cliente"
        
        mensagem = (
            f"🔔 *Lembrete de Agendamento*\n\n"
            f"Olá, {primeiro_nome}! 👋\n\n"
            f"Este é um lembrete do seu agendamento *amanhã*:\n\n"
            f"📅 Data: *{data}*\n"
            f"⏰ Horário: *{hora}*\n\n"
            f"💈 Te esperamos na Barbearia Veinho Corts!\n\n"
            f"⚠️ Caso precise cancelar ou remarcar, digite *menu* e escolha a opção 1."
        )
        
        # Enviar mensagem via WAHA
        if _waha_send_function:
            _waha_send_function(chat_id, mensagem)
        
        # Atualizar timestamp do lembrete
        _atualizar_lembrete_enviado(ag)
        
        logger.info(f"📤 Lembrete 1 dia enviado: {nome} - {data} {hora}")
        return True
    
    except Exception as e:
        logger.error(f"Erro ao enviar lembrete 1 dia: {e}")
        return False


def _enviar_lembrete_2_horas(ag: Dict[str, Any]) -> bool:
    """Envia lembrete de 2 horas antes."""
    try:
        chat_id = ag.get("ChatId", "")
        nome = ag.get("ClienteNome", "Cliente")
        data = ag.get("Data", "")
        hora = ag.get("Hora", "")
        
        if not chat_id:
            return False
        
        # Formatar nome (só primeiro nome)
        primeiro_nome = nome.split()[0] if nome else "Cliente"
        
        mensagem = (
            f"⏰ *Lembrete Urgente!*\n\n"
            f"Olá, {primeiro_nome}! 👋\n\n"
            f"Seu agendamento é *daqui a 2 horas*:\n\n"
            f"📅 Data: *{data}*\n"
            f"⏰ Horário: *{hora}*\n\n"
            f"💈 Estamos te esperando!\n\n"
            f"🚨 Caso precise cancelar, faça o quanto antes para liberarmos o horário."
        )
        
        # Enviar mensagem via WAHA
        if _waha_send_function:
            _waha_send_function(chat_id, mensagem)
        
        # Atualizar timestamp do lembrete
        _atualizar_lembrete_enviado(ag)
        
        logger.info(f"📤 Lembrete 2h enviado: {nome} - {data} {hora}")
        return True
    
    except Exception as e:
        logger.error(f"Erro ao enviar lembrete 2h: {e}")
        return False


def _atualizar_lembrete_enviado(ag: Dict[str, Any]):
    """Atualiza o campo LembreteEnviado na planilha."""
    from openpyxl import load_workbook
    
    try:
        chave = ag.get("Chave", "")
        if not chave:
            return
        
        wb = load_workbook(es.FILE_PATH)
        ws = wb[es.SHEET_AG]
        
        es._ensure_headers(ws, es.HEADERS_AG)
        hm = es._get_header_map(ws)
        
        c_lembrete = hm.get("LembreteEnviado")
        
        if not c_lembrete:
            return
        
        # Buscar linha pela chave
        row = es._find_row_by_key(ws, chave)
        
        if row:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row=row, column=c_lembrete, value=timestamp)
            wb.save(es.FILE_PATH)
    
    except Exception as e:
        logger.error(f"Erro ao atualizar lembrete enviado: {e}")


# Função de teste manual
def testar_lembretes_manual():
    """Função para testar lembretes manualmente (chamar via admin ou debug)."""
    logger.info("🧪 TESTE MANUAL: Verificando lembretes...")
    enviar_lembretes_1_dia()
    enviar_lembretes_2_horas()
    logger.info("🧪 TESTE MANUAL: Concluído")
