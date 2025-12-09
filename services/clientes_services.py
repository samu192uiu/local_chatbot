# services/clientes_services.py
import os
import re
import hashlib
from datetime import datetime
from openpyxl import Workbook, load_workbook

# =========================
# Config
# =========================
FILE_PATH = os.getenv("CLIENTES_XLSX", "/app/data/clientes.xlsx")
SALT = os.getenv("PIN_SALT", "barbearia_salt")

# Cabeçalhos canonizados (ordem importa no XLSX)
HEADERS = [
    "ID", "CPF", "Nome", "Nascimento", "Telefone", "Email",
    "ChatId", "PinHash", "UltimoLogin", "CriadoEm", "AtualizadoEm",
    "TentativasPin", "BloqueadoAte"  # controle de tentativas de login
]

# =========================
# Utils
# =========================
def _cpf_puro(cpf: str) -> str:
    return re.sub(r"\D", "", cpf or "")

def _phone_puro(phone: str) -> str:
    return re.sub(r"\D", "", phone or "")

def _jid_to_phone(jid: str) -> str:
    # "5511999999999@c.us" -> "5511999999999"
    return _phone_puro((jid or "").split("@")[0])

def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _hash_pin(pin: str) -> str:
    return hashlib.sha256(f"{SALT}:{pin}".encode("utf-8")).hexdigest()

def _open_ws():
    init_planilha()  # garante arquivo/cabeçalho
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    return wb, ws

def _col_index(col_name: str) -> int:
    return HEADERS.index(col_name) + 1

def _row_to_dict(ws, row_idx: int) -> dict:
    data = {}
    for col, h in enumerate(HEADERS, 1):
        data[h] = ws.cell(row=row_idx, column=col).value
    return data

def _find_row_by(ws, col_name: str, value: str) -> int | None:
    col_idx = _col_index(col_name)
    value = str(value or "").strip()
    for r in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=col_idx).value
        if str(cell_val or "").strip() == value:
            return r
    return None

def _next_id(ws) -> int:
    """Retorna o próximo ID inteiro com base no maior ID já existente."""
    col = _col_index("ID")
    max_id = 0
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=col).value
        try:
            max_id = max(max_id, int(val))
        except Exception:
            pass
    return (max_id + 1) if max_id >= 0 else 1

def _ensure_headers(ws) -> bool:
    """Garante cabeçalho completo na ordem. Retorna True se alterou algo."""
    changed = False
    # Cabeçalho vazio? Escreve tudo.
    empty_header = True
    for c in range(1, len(HEADERS) + 1):
        if ws.cell(row=1, column=c).value:
            empty_header = False
            break
    if empty_header:
        for col, h in enumerate(HEADERS, 1):
            ws.cell(row=1, column=col, value=h)
        return True

    # Força nomes na posição correta (migração leve)
    for col, h in enumerate(HEADERS, 1):
        if ws.cell(row=1, column=col).value != h:
            ws.cell(row=1, column=col, value=h)
            changed = True
    return changed

def _ensure_row_has_id(ws, row_idx: int) -> bool:
    """Se a linha não tem ID, atribui o próximo. Retorna True se alterou."""
    id_cell = ws.cell(row=row_idx, column=_col_index("ID"))
    if id_cell.value in (None, "", 0):
        id_cell.value = _next_id(ws)
        return True
    return False

# =========================
# Init / Migração leve
# =========================
def init_planilha():
    os.makedirs(os.path.dirname(FILE_PATH), exist_ok=True)
    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        for col, h in enumerate(HEADERS, 1):
            ws.cell(row=1, column=col, value=h)
        wb.save(FILE_PATH)
        return

    # Arquivo existe: garantir cabeçalho
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    changed = _ensure_headers(ws)
    if changed:
        wb.save(FILE_PATH)

# =========================
# CRUD / Upserts
# =========================
def get_by_id(id_val: int | str) -> dict | None:
    _, ws = _open_ws()
    r = _find_row_by(ws, "ID", str(id_val))
    return _row_to_dict(ws, r) if r else None

def get_by_cpf(cpf: str) -> dict | None:
    cpf = _cpf_puro(cpf)
    _, ws = _open_ws()
    r = _find_row_by(ws, "CPF", cpf)
    return _row_to_dict(ws, r) if r else None

def get_by_chat_id(chat_id: str) -> dict | None:
    _, ws = _open_ws()
    r = _find_row_by(ws, "ChatId", chat_id or "")
    return _row_to_dict(ws, r) if r else None

def get_by_phone(phone: str) -> dict | None:
    phone = _phone_puro(phone)
    _, ws = _open_ws()
    r = _find_row_by(ws, "Telefone", phone)
    return _row_to_dict(ws, r) if r else None

def find_by_chat(chat_id: str) -> dict | None:
    """Fallback por telefone do JID."""
    phone = _jid_to_phone(chat_id)
    return get_by_phone(phone)

def exists_cpf(cpf: str) -> bool:
    return get_by_cpf(cpf) is not None

def create_or_update_client(rec: dict) -> dict:
    """
    Upsert por CPF. Se não existir, cria com ID novo; se existir, atualiza campos.
    Campos aceitos: CPF, Nome, Nascimento, Telefone, Email, ChatId
    """
    cpf = _cpf_puro(rec.get("CPF"))
    if not cpf:
        raise ValueError("CPF é obrigatório em create_or_update_client")

    wb, ws = _open_ws()
    r = _find_row_by(ws, "CPF", cpf)
    now = _now_str()

    if not r:
        r = ws.max_row + 1
        ws.cell(row=r, column=_col_index("CPF"), value=cpf)
        ws.cell(row=r, column=_col_index("CriadoEm"), value=now)
        _ensure_row_has_id(ws, r)

    # Normaliza telefone
    tel = rec.get("Telefone")
    if tel:
        rec["Telefone"] = _phone_puro(str(tel))

    # Preenche campos simples
    for k in ("Nome", "Nascimento", "Telefone", "Email", "ChatId"):
        if k in rec:
            ws.cell(row=r, column=_col_index(k), value=rec.get(k))

    # Timestamps
    ws.cell(row=r, column=_col_index("AtualizadoEm"), value=now)

    wb.save(FILE_PATH)
    out = _row_to_dict(ws, r)
    return out

# Alias comum
def upsert_client(rec: dict) -> dict:
    return create_or_update_client(rec)

# =========================
# PIN / Login
# =========================
def set_pin_for_cpf(cpf: str, pin: str) -> bool:
    cpf = _cpf_puro(cpf)
    wb, ws = _open_ws()
    r = _find_row_by(ws, "CPF", cpf)
    if not r:
        return False
    ws.cell(row=r, column=_col_index("PinHash"), value=_hash_pin(pin))
    wb.save(FILE_PATH)  # <-- FIX: era ws.save, agora wb.save
    return True

def set_pin(cpf: str, pin: str) -> bool:
    return set_pin_for_cpf(cpf, pin)

def verify_pin(cpf: str, pin: str) -> bool:
    cpf = _cpf_puro(cpf)
    _, ws = _open_ws()
    r = _find_row_by(ws, "CPF", cpf)
    if not r:
        return False
    saved = ws.cell(row=r, column=_col_index("PinHash")).value or ""
    return saved == _hash_pin(pin)

def incrementar_tentativa_pin(cpf: str) -> int:
    """
    Incrementa contador de tentativas de PIN e retorna o total.
    Se atingir 3 tentativas, bloqueia por 15 minutos.
    
    Returns:
        Número atual de tentativas
    """
    cpf = _cpf_puro(cpf)
    wb, ws = _open_ws()
    r = _find_row_by(ws, "CPF", cpf)
    if not r:
        return 0
    
    tentativas = int(ws.cell(row=r, column=_col_index("TentativasPin")).value or 0)
    tentativas += 1
    
    ws.cell(row=r, column=_col_index("TentativasPin"), value=tentativas)
    
    # Se atingiu 3 tentativas, bloqueia por 15 minutos
    if tentativas >= 3:
        from datetime import datetime, timedelta
        bloqueado_ate = (datetime.now() + timedelta(minutes=15)).strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=r, column=_col_index("BloqueadoAte"), value=bloqueado_ate)
    
    wb.save(FILE_PATH)
    return tentativas

def esta_bloqueado(cpf: str) -> bool:
    """
    Verifica se o CPF está temporariamente bloqueado por tentativas de PIN.
    
    Returns:
        True se bloqueado, False caso contrário
    """
    cpf = _cpf_puro(cpf)
    _, ws = _open_ws()
    r = _find_row_by(ws, "CPF", cpf)
    if not r:
        return False
    
    bloqueado_ate_str = ws.cell(row=r, column=_col_index("BloqueadoAte")).value or ""
    if not bloqueado_ate_str:
        return False
    
    try:
        from datetime import datetime
        bloqueado_ate = datetime.strptime(str(bloqueado_ate_str), "%Y-%m-%d %H:%M:%S")
        return datetime.now() < bloqueado_ate
    except Exception:
        return False

def resetar_tentativas_pin(cpf: str) -> None:
    """
    Reseta contador de tentativas e remove bloqueio (chamado após login bem-sucedido).
    """
    cpf = _cpf_puro(cpf)
    wb, ws = _open_ws()
    r = _find_row_by(ws, "CPF", cpf)
    if not r:
        return
    
    ws.cell(row=r, column=_col_index("TentativasPin"), value=0)
    ws.cell(row=r, column=_col_index("BloqueadoAte"), value="")
    wb.save(FILE_PATH)

def touch_login(cpf: str) -> None:
    cpf = _cpf_puro(cpf)
    wb, ws = _open_ws()
    r = _find_row_by(ws, "CPF", cpf)
    if not r:
        return
    now = _now_str()
    ws.cell(row=r, column=_col_index("UltimoLogin"), value=now)
    ws.cell(row=r, column=_col_index("AtualizadoEm"), value=now)
    # Resetar tentativas ao fazer login com sucesso
    ws.cell(row=r, column=_col_index("TentativasPin"), value=0)
    ws.cell(row=r, column=_col_index("BloqueadoAte"), value="")
    wb.save(FILE_PATH)

# =========================
# Listagem / Busca
# =========================
def count_clients() -> int:
    _, ws = _open_ws()
    return max(0, ws.max_row - 1)

def list_all_clients(offset: int = 0, limit: int = 50) -> list[dict]:
    _, ws = _open_ws()
    start = 2 + max(0, offset)
    end = min(ws.max_row, start + max(1, limit) - 1)
    out = []
    for r in range(start, end + 1):
        out.append(_row_to_dict(ws, r))
    return out

def search_clients(query: str) -> list[dict]:
    q = (query or "").strip().lower()
    if not q:
        return []
    _, ws = _open_ws()
    results = []
    for r in range(2, ws.max_row + 1):
        rec = _row_to_dict(ws, r)
        hay = " ".join(str(rec.get(k) or "") for k in ("Nome","CPF","Telefone","Email","ChatId")).lower()
        if q in hay:
            results.append(rec)
    return results

def list_logins_links() -> list[dict]:
    """
    Retorna vínculos de login no formato consumido pelo painel admin:
    [{"chat_id": "...@c.us", "cpf": "99999999999", "nome": "Fulano", "id": 12}]
    """
    out = []
    for rec in list_all_clients(0, 100000):
        chat = str(rec.get("ChatId") or "").strip()
        cpf  = _cpf_puro(rec.get("CPF") or "")
        nome = str(rec.get("Nome") or "").strip()
        rid  = rec.get("ID")
        if chat and cpf:
            out.append({"chat_id": chat, "cpf": cpf, "nome": nome, "id": rid})
    return out

def _auth_list_links() -> list[dict]:
    # compat com painel admin que procura esse helper
    return list_logins_links()

# =========================
# Sanidade / Migração suave
# =========================
def sanity_fix_and_report() -> dict:
    """
    Corrige linhas sem ID e garante cabeçalho. Retorna um resumo útil para debug.
    """
    wb, ws = _open_ws()
    fixed_ids = 0
    for r in range(2, ws.max_row + 1):
        if _ensure_row_has_id(ws, r):
            fixed_ids += 1
    if fixed_ids:
        wb.save(FILE_PATH)

    total = count_clients()
    sample = list_all_clients(0, min(total, 5))
    return {
        "arquivo": FILE_PATH,
        "total_clientes": total,
        "ids_corrigidos": fixed_ids,
        "amostra": sample
    }
