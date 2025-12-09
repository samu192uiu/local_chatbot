# services/excel_services.py
import os
import json
import threading
from datetime import datetime, timedelta
from typing import Iterable, Dict, List, Optional, Tuple, Any
from openpyxl import Workbook, load_workbook
import logging

logger = logging.getLogger("ZapWaha")

# =========================================================
# Config e Lock Global
# =========================================================
FILE_PATH = os.getenv("AGENDAMENTOS_XLSX", "/app/data/agendamentos.xlsx")
SHEET_AG = os.getenv("AG_SHEET_NAME", "Agendamentos")
SHEET_CLIENTES = os.getenv("CLIENTES_SHEET_NAME", "Clientes")  # opcional (para _read_rows_clientes)
FERIADOS_JSON = os.path.join(os.path.dirname(__file__), "..", "config", "feriados.json")

# Lock global para operações críticas de reserva
_RESERVA_LOCK = threading.Lock()

HEADERS_AG = [
    "Chave",         # chave única do lançamento
    "Data",          # DD/MM/AAAA
    "Hora",          # HH:MM
    "ChatId",        # jid/whatsapp do cliente
    "ClienteID",     # << NOVO: ID único vindo do clientes_services
    "ClienteNome",
    "Nascimento",
    "CPF",
    "ServicoID",     # << NOVO: ID do serviço (corte_simples, platinado, etc.)
    "ServicoDuracao", # << NOVO: Duração em minutos
    "Status",        # Reservado | Confirmado | Expirado | Cancelado | Remarcado
    "ReservadoEm",   # << NOVO: Timestamp da reserva
    "ReservadoAte",  # << NOVO: Timestamp de expiração da reserva
    "PagamentoID",   # << NOVO: ID do pagamento (Mercado Pago - futuro)
    "PagamentoStatus", # << NOVO: pending/approved/rejected (futuro)
    "ValorPago",
    "CriadoEm",      # timestamp
    "Remarcacoes",   # contador de remarcações (máximo 1)
    "LembreteEnviado", # timestamp do último lembrete enviado
]

# estados que bloqueiam o mesmo slot (Data+Hora)
BLOCKING_STATUSES = set([
    "Reservado",         # Reserva temporária ativa
    "Pendente Pagamento",
    "Confirmado",
    "Reagendado",
    "Remarcado",
])

# =========================================================
# Helpers de planilha
# =========================================================
def _ensure_file_and_sheet():
    """Cria arquivo/sheet se não existirem; garante cabeçalho da aba de Agendamentos."""
    os.makedirs(os.path.dirname(FILE_PATH), exist_ok=True)
    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_AG
        for idx, h in enumerate(HEADERS_AG, 1):
            ws.cell(row=1, column=idx, value=h)
        wb.save(FILE_PATH)
        return

    wb = load_workbook(FILE_PATH)
    if SHEET_AG not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_AG)
        for idx, h in enumerate(HEADERS_AG, 1):
            ws.cell(row=1, column=idx, value=h)
        wb.save(FILE_PATH)
        return

    # garantir cabeçalhos na aba
    ws = wb[SHEET_AG]
    _ensure_headers(ws, HEADERS_AG)
    wb.save(FILE_PATH)

def _open_ws():
    _ensure_file_and_sheet()
    wb = load_workbook(FILE_PATH)
    ws = wb[SHEET_AG]
    return wb, ws

def _get_header_map(ws) -> dict:
    """nome_coluna -> índice (1-based)"""
    m = {}
    for c in range(1, ws.max_column + 1):
        name = str(ws.cell(row=1, column=c).value or "").strip()
        if name:
            m[name] = c
    return m

def _ensure_headers(ws, headers: List[str]):
    """Garante que todas as colunas de headers existam; se faltarem, adiciona ao final."""
    header_map = _get_header_map(ws)
    changed = False
    for h in headers:
        if h not in header_map:
            new_col = ws.max_column + 1 if ws.max_column >= 1 else 1
            ws.cell(row=1, column=new_col, value=h)
            header_map[h] = new_col
            changed = True
    return changed

def _row_to_dict(ws, row_idx: int) -> dict:
    hm = _get_header_map(ws)
    out = {}
    for h in HEADERS_AG:
        col = hm.get(h)
        out[h] = ws.cell(row=row_idx, column=col).value if col else None
    return out

def _make_row(
    ws,
    chave: str,
    data_str: str,
    hora_str: str,
    chat_id: str,
    cliente_id: Optional[str],
    cliente_nome: Optional[str],
    nasc: Optional[str],
    cpf: Optional[str],
    status: str,
    valor_pago: Optional[Any],
    servico_id: Optional[str] = None,
    servico_duracao: Optional[int] = None,
    reservado_em: Optional[str] = None,
    reservado_ate: Optional[str] = None,
):
    hm = _get_header_map(ws)
    r = ws.max_row + 1
    def setv(colname, val):
        col = hm.get(colname)
        if col:
            ws.cell(row=r, column=col, value=val)

    setv("Chave", chave)
    setv("Data", data_str)
    setv("Hora", hora_str)
    setv("ChatId", chat_id)
    setv("ClienteID", cliente_id or "")
    setv("ClienteNome", cliente_nome or "")
    setv("Nascimento", nasc or "")
    setv("CPF", cpf or "")
    setv("ServicoID", servico_id or "corte_simples")
    setv("ServicoDuracao", servico_duracao or 40)
    setv("Status", status or "")
    setv("ReservadoEm", reservado_em or "")
    setv("ReservadoAte", reservado_ate or "")
    setv("PagamentoID", "")
    setv("PagamentoStatus", "")
    setv("ValorPago", valor_pago)
    setv("CriadoEm", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    return r

# =========================================================
# API pública usada pelo fluxo
# =========================================================
def make_key(data_str: str, hora_str: str, chat_id: str) -> str:
    """Gera uma chave canônica a partir do trio (data, hora, chat)."""
    return f"{data_str}|{hora_str}|{chat_id}"

def verificar_disponibilidade(data_str: str, hora_str: str, servico_id: Optional[str] = None) -> bool:
    """
    Verifica se um horário está disponível considerando serviços fracionados.
    
    Para serviços simples: verifica se há conflito no slot Data+Hora.
    Para serviços fracionados: verifica se os períodos ocupados do serviço
    conflitam com períodos ocupados de outros agendamentos.
    
    Args:
        data_str: Data no formato DD/MM/YYYY
        hora_str: Hora no formato HH:MM
        servico_id: ID do serviço (ex: 'platinado', 'corte_simples')
    
    Returns:
        True se disponível, False caso contrário
    """
    # Tentar usar verificação fracionada se disponível
    try:
        from services import servicos_fracionados as sf
        
        # Se servico_id não fornecido, assume simples
        if not servico_id:
            servico_id = "corte_simples"
        
        # Buscar todos os agendamentos ativos
        _, ws = _open_ws()
        hm = _get_header_map(ws)
        c_data = hm.get("Data")
        c_hora = hm.get("Hora")
        c_status = hm.get("Status")
        c_servico = hm.get("ServicoID")
        
        agendamentos_existentes = []
        for r in range(2, ws.max_row + 1):
            st = str(ws.cell(row=r, column=c_status).value or "").strip()
            if st not in BLOCKING_STATUSES:
                continue
            
            d = str(ws.cell(row=r, column=c_data).value or "").strip()
            h = str(ws.cell(row=r, column=c_hora).value or "").strip()
            s = str(ws.cell(row=r, column=c_servico).value or "corte_simples").strip()
            
            if d and h:
                agendamentos_existentes.append({
                    "Data": d,
                    "Hora": h,
                    "ServicoID": s
                })
        
        # Usar verificação fracionada
        disponivel, mensagem = sf.verificar_disponibilidade_fracionado(
            servico_id=servico_id,
            data=data_str,
            horario_inicio=hora_str,
            agendamentos_existentes=agendamentos_existentes
        )
        
        return disponivel
        
    except Exception as e:
        # Fallback para verificação simples
        _, ws = _open_ws()
        hm = _get_header_map(ws)
        c_data = hm.get("Data")
        c_hora = hm.get("Hora")
        c_status = hm.get("Status")

        for r in range(2, ws.max_row + 1):
            d = str(ws.cell(row=r, column=c_data).value or "").strip()
            h = str(ws.cell(row=r, column=c_hora).value or "").strip()
            if d == data_str and h == hora_str:
                st = str(ws.cell(row=r, column=c_status).value or "").strip()
                if st in BLOCKING_STATUSES:
                    return False
        return True

def adicionar_agendamento(
    data_str: str,
    hora_str: str,
    chat_id: str,
    status: str = "Pendente Pagamento",
    cliente_nome: Optional[str] = None,
    data_nasc: Optional[str] = None,
    cpf: Optional[str] = None,
    valor_pago: Optional[Any] = None,
    cliente_id: Optional[str] = None,
    servico_id: Optional[str] = None,
    servico_duracao: Optional[int] = None,
) -> str:
    """
    Cria linha na planilha de agendamentos e retorna a 'Chave'.
    Se já houver bloqueio no slot, levanta ValueError.
    """
    if not verificar_disponibilidade(data_str, hora_str, servico_id):
        raise ValueError("Horário indisponível")

    wb, ws = _open_ws()
    _ensure_headers(ws, HEADERS_AG)
    chave = make_key(data_str, hora_str, chat_id)

    _make_row(
        ws=ws,
        chave=chave,
        data_str=data_str,
        hora_str=hora_str,
        chat_id=chat_id,
        cliente_id=cliente_id,
        cliente_nome=cliente_nome,
        nasc=data_nasc,
        cpf=cpf,
        status=status,
        valor_pago=valor_pago,
        servico_id=servico_id,
        servico_duracao=servico_duracao,
    )
    wb.save(FILE_PATH)
    return chave

def _find_row_by_key(ws, chave: str) -> Optional[int]:
    hm = _get_header_map(ws)
    c_key = hm.get("Chave")
    if not c_key:
        return None
    chave = str(chave or "").strip()
    for r in range(2, ws.max_row + 1):
        v = str(ws.cell(row=r, column=c_key).value or "").strip()
        if v == chave:
            return r
    return None

def _find_row_by_triplet(ws, data_str: str, hora_str: str, chat_id: str) -> Optional[int]:
    hm = _get_header_map(ws)
    c_data = hm.get("Data"); c_hora = hm.get("Hora"); c_chat = hm.get("ChatId")
    if not (c_data and c_hora and c_chat):
        return None
    ds = str(data_str or "").strip()
    hs = str(hora_str or "").strip()
    jid = str(chat_id or "").strip()
    for r in range(2, ws.max_row + 1):
        d = str(ws.cell(row=r, column=c_data).value or "").strip()
        h = str(ws.cell(row=r, column=c_hora).value or "").strip()
        j = str(ws.cell(row=r, column=c_chat).value or "").strip()
        if d == ds and h == hs and j == jid:
            return r
    return None

def atualizar_status_por_chave(*args) -> bool:
    """
    Modo 1 (recomendado): atualizar_status_por_chave(chave, novo_status)
    Modo 2 (compat):      atualizar_status_por_chave(data, hora, chat_id, novo_status)
    """
    wb, ws = _open_ws()
    _ensure_headers(ws, HEADERS_AG)
    hm = _get_header_map(ws)
    c_status = hm.get("Status")

    row = None
    new_status = None

    if len(args) == 2:
        chave, new_status = args
        row = _find_row_by_key(ws, chave)
    elif len(args) == 4:
        data_str, hora_str, chat_id, new_status = args
        row = _find_row_by_triplet(ws, data_str, hora_str, chat_id)
    else:
        raise TypeError("Uso: atualizar_status_por_chave(chave, status) OU atualizar_status_por_chave(data,hora,chat_id,status)")

    if row is None:
        return False

    ws.cell(row=row, column=c_status, value=str(new_status or "").strip())
    wb.save(FILE_PATH)
    return True

def atualizar_status(chave: str, new_status: str) -> bool:
    """Fallback simples por chave."""
    try:
        return atualizar_status_por_chave(chave, new_status)
    except Exception:
        return False

def atualizar_agendamento_remarcar(chave_antiga: str, nova_data: str, nova_hora: str) -> Tuple[bool, Optional[str]]:
    """
    Atualiza um agendamento existente com nova data/hora (remarcação).
    
    Args:
        chave_antiga: Chave do agendamento original
        nova_data: Nova data no formato DD/MM/YYYY
        nova_hora: Novo horário no formato HH:MM
    
    Returns:
        (sucesso, mensagem_erro)
        - sucesso: True se atualizado, False se erro
        - mensagem_erro: None se sucesso, string com erro caso contrário
    """
    wb, ws = _open_ws()
    _ensure_headers(ws, HEADERS_AG)
    hm = _get_header_map(ws)
    
    row = _find_row_by_key(ws, chave_antiga)
    if row is None:
        return False, "Agendamento não encontrado."
    
    # Verificar contador de remarcações
    c_remarcacoes = hm.get("Remarcacoes")
    remarcacoes_atuais = 0
    if c_remarcacoes:
        valor = ws.cell(row=row, column=c_remarcacoes).value
        try:
            remarcacoes_atuais = int(valor) if valor else 0
        except (ValueError, TypeError):
            remarcacoes_atuais = 0
    
    # Verificar se já atingiu o limite de remarcações
    if remarcacoes_atuais >= 1:
        return False, "limite_atingido"
    
    # Obter ChatId para gerar nova chave
    c_chat = hm.get("ChatId")
    chat_id = str(ws.cell(row=row, column=c_chat).value or "").strip()
    
    # Gerar nova chave
    nova_chave = make_key(nova_data, nova_hora, chat_id)
    
    # Atualizar campos
    c_chave = hm.get("Chave")
    c_data = hm.get("Data")
    c_hora = hm.get("Hora")
    c_status = hm.get("Status")
    
    if c_chave:
        ws.cell(row=row, column=c_chave, value=nova_chave)
    if c_data:
        ws.cell(row=row, column=c_data, value=nova_data)
    if c_hora:
        ws.cell(row=row, column=c_hora, value=nova_hora)
    if c_status:
        ws.cell(row=row, column=c_status, value="Confirmado")
    if c_remarcacoes:
        ws.cell(row=row, column=c_remarcacoes, value=remarcacoes_atuais + 1)
    
    wb.save(FILE_PATH)
    return True, None

# =========================================================
# Leitura (para painéis/relatórios)
# =========================================================
def _read_rows(sheet: str = SHEET_AG) -> Iterable[Dict[str, Any]]:
    """
    Itera linhas como dicionário (apenas colunas conhecidas do respectivo sheet).
    Para Agendamentos, usa HEADERS_AG; para outros, devolve todas as colunas encontradas.
    """
    _ensure_file_and_sheet()
    wb = load_workbook(FILE_PATH)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    header_map = {}
    headers = []

    # detecta cabeçalho
    for c in range(1, ws.max_column + 1):
        name = str(ws.cell(row=1, column=c).value or "").strip()
        if name:
            header_map[name] = c
            headers.append(name)

    def row_to_dict_row(r: int) -> Dict[str, Any]:
        d = {}
        for h in headers:
            col = header_map[h]
            d[h] = ws.cell(row=r, column=col).value
        return d

    out = []
    for r in range(2, ws.max_row + 1):
        # ignora linhas totalmente vazias
        empty = True
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value not in (None, ""):
                empty = False
                break
        if empty:
            continue
        out.append(row_to_dict_row(r))
    return out

def _read_rows_clientes() -> Iterable[Dict[str, Any]]:
    """
    Opcional: se você mantiver a aba 'Clientes' dentro deste MESMO arquivo,
    esta função ajuda o painel admin a listar vínculos.
    Cabeçalhos esperados: ['CPF','Nome','Nascimento','Telefone','Email','ChatId','PinHash','UltimoLogin','ClienteID']
    """
    _ensure_file_and_sheet()
    wb = load_workbook(FILE_PATH)
    if SHEET_CLIENTES not in wb.sheetnames:
        return []  # se você guarda clientes em outro arquivo (clientes.xlsx), tudo bem
    ws = wb[SHEET_CLIENTES]

    # coleta cabeçalhos dinamicamente
    header_map = {}
    headers = []
    for c in range(1, ws.max_column + 1):
        name = str(ws.cell(row=1, column=c).value or "").strip()
        if name:
            header_map[name] = c
            headers.append(name)

    out = []
    for r in range(2, ws.max_row + 1):
        d = {}
        for h in headers:
            col = header_map[h]
            d[h] = ws.cell(row=r, column=col).value
        out.append(d)
    return out

def tem_agendamento_ativo_na_semana(chat_id: str) -> Tuple[bool, Optional[Dict[str, Any]]]:
    """
    Verifica se o chat_id já possui um agendamento ativo (futuro ou de hoje).
    
    Retorna:
        (tem_agendamento, dados_do_agendamento)
        - tem_agendamento: True se existe agendamento ativo
        - dados_do_agendamento: Dict com Data, Hora, Status se existir, None caso contrário
    """
    wb, ws = _open_ws()
    hm = _get_header_map(ws)
    
    c_chat = hm.get("ChatId")
    c_data = hm.get("Data")
    c_hora = hm.get("Hora")
    c_status = hm.get("Status")
    
    if not all([c_chat, c_data, c_hora, c_status]):
        return False, None
    
    agora = datetime.now()
    hoje_str = agora.strftime("%d/%m/%Y")
    
    for r in range(2, ws.max_row + 1):
        jid = str(ws.cell(row=r, column=c_chat).value or "").strip()
        status = str(ws.cell(row=r, column=c_status).value or "").strip()
        
        # Filtrar apenas agendamentos deste cliente
        if jid != chat_id:
            continue
        
        # Verificar se tem status bloqueante (não cancelado)
        if status not in BLOCKING_STATUSES:
            continue
        
        data_str = str(ws.cell(row=r, column=c_data).value or "").strip()
        hora_str = str(ws.cell(row=r, column=c_hora).value or "").strip()
        
        try:
            # Parsear data e hora
            data_hora = datetime.strptime(f"{data_str} {hora_str}", "%d/%m/%Y %H:%M")
            
            # Verificar se é hoje ou futuro
            if data_str == hoje_str or data_hora >= agora:
                agendamento_info = {
                    "Data": data_str,
                    "Hora": hora_str,
                    "Status": status,
                    "data_hora_obj": data_hora
                }
                return True, agendamento_info
        except Exception:
            continue
    
    return False, None

def buscar_proximo_agendamento(chat_id: str) -> Optional[Dict[str, Any]]:
    """
    Busca o próximo agendamento confirmado do cliente (por chat_id).
    Retorna dict com as informações ou None se não houver.
    Considera agendamentos futuros E agendamentos de hoje (mesmo que já tenha passado a hora).
    """
    wb, ws = _open_ws()
    hm = _get_header_map(ws)
    
    c_chat = hm.get("ChatId")
    c_data = hm.get("Data")
    c_hora = hm.get("Hora")
    c_status = hm.get("Status")
    
    if not all([c_chat, c_data, c_hora, c_status]):
        return None
    
    agora = datetime.now()
    hoje_str = agora.strftime("%d/%m/%Y")
    agendamentos_validos = []
    
    for r in range(2, ws.max_row + 1):
        jid = str(ws.cell(row=r, column=c_chat).value or "").strip()
        status = str(ws.cell(row=r, column=c_status).value or "").strip().lower()
        
        # Filtrar apenas agendamentos deste cliente e confirmados/pendentes
        if jid != chat_id or status not in ["confirmado", "pendente pagamento"]:
            continue
        
        data_str = str(ws.cell(row=r, column=c_data).value or "").strip()
        hora_str = str(ws.cell(row=r, column=c_hora).value or "").strip()
        
        try:
            # Parsear data e hora
            data_hora = datetime.strptime(f"{data_str} {hora_str}", "%d/%m/%Y %H:%M")
            
            # Incluir agendamentos de hoje OU futuros
            # (mesmo que o horário de hoje já tenha passado, mostra)
            if data_str == hoje_str or data_hora > agora:
                agendamento = _row_to_dict(ws, r)
                agendamento['data_hora_obj'] = data_hora
                agendamentos_validos.append(agendamento)
        except Exception:
            continue
    
    # Retornar o mais próximo (ordenar por data/hora)
    if agendamentos_validos:
        agendamentos_validos.sort(key=lambda x: x['data_hora_obj'])
        return agendamentos_validos[0]
    
    return None


def buscar_historico_completo(cpf: str) -> List[Dict[str, Any]]:
    """
    Retorna TODOS os agendamentos de um cliente (passados e futuros, qualquer status).
    
    Args:
        cpf: CPF do cliente (apenas números)
    
    Returns:
        Lista de dicts com agendamentos ordenados por data (mais recente primeiro)
        Cada dict contém: Chave, Data, Hora, Status, ClienteNome, ValorPago, CriadoEm, etc.
    """
    wb, ws = _open_ws()
    _ensure_headers(ws, HEADERS_AG)
    hm = _get_header_map(ws)
    
    c_cpf = hm.get("CPF")
    c_data = hm.get("Data")
    c_hora = hm.get("Hora")
    
    if not all([c_cpf, c_data, c_hora]):
        return []
    
    cpf_limpo = cpf.strip()
    agendamentos = []
    
    for r in range(2, ws.max_row + 1):
        cpf_row = str(ws.cell(row=r, column=c_cpf).value or "").strip()
        
        if cpf_row == cpf_limpo:
            agendamento = _row_to_dict(ws, r)
            
            # Adicionar objeto datetime para ordenação
            data_str = agendamento.get("Data", "")
            hora_str = agendamento.get("Hora", "")
            try:
                agendamento['data_hora_obj'] = datetime.strptime(
                    f"{data_str} {hora_str}", "%d/%m/%Y %H:%M"
                )
            except Exception:
                agendamento['data_hora_obj'] = datetime.min
            
            agendamentos.append(agendamento)
    
    # Ordenar por data (mais recente primeiro)
    agendamentos.sort(key=lambda x: x['data_hora_obj'], reverse=True)
    
    return agendamentos


def eh_feriado(data_str: str) -> bool:
    """
    Verifica se a data é um feriado bloqueado.
    
    Args:
        data_str: Data no formato DD/MM/YYYY
    
    Returns:
        True se é feriado, False caso contrário
    """
    try:
        if not os.path.exists(FERIADOS_JSON):
            return False
        
        with open(FERIADOS_JSON, 'r', encoding='utf-8') as f:
            config = json.load(f)
            feriados = config.get("feriados", [])
            return data_str in feriados
    except Exception:
        return False


def horario_muito_proximo(data_str: str, hora_str: str, horas_minimas: int = 2) -> bool:
    """
    Verifica se o horário está muito próximo do momento atual.
    
    Args:
        data_str: Data no formato DD/MM/YYYY
        hora_str: Hora no formato HH:MM
        horas_minimas: Número mínimo de horas de antecedência (padrão: 2)
    
    Returns:
        True se faltam menos de X horas para o horário, False caso contrário
    """
    try:
        from datetime import timedelta
        data_hora = datetime.strptime(f"{data_str} {hora_str}", "%d/%m/%Y %H:%M")
        agora = datetime.now()
        diferenca = data_hora - agora
        
        # Se já passou ou falta menos de X horas
        return diferenca.total_seconds() < (horas_minimas * 3600)
    except Exception:
        return False



# =========================================================
# Sistema de Reserva Temporária (Anti-Conflito)
# =========================================================

DURACAO_RESERVA_MINUTOS = 10  # Tempo para confirmar agendamento

def reservar_slot_temporario(
    data_str: str,
    hora_str: str,
    chat_id: str,
    servico_id: str,
    servico_duracao: int,
    cliente_nome: Optional[str] = None,
    cliente_id: Optional[str] = None,
    duracao_reserva_min: int = DURACAO_RESERVA_MINUTOS
) -> Dict[str, Any]:
    """
    Reserva um slot temporariamente por X minutos.
    
    Args:
        data_str: Data DD/MM/YYYY
        hora_str: Hora HH:MM
        chat_id: WhatsApp ID
        servico_id: ID do serviço
        servico_duracao: Duração do serviço em minutos
        duracao_reserva_min: Minutos de reserva (padrão: 10)
    
    Returns:
        Dict com: sucesso, chave, expira_em, mensagem
    """
    # LOCK CRÍTICO: verificar e reservar devem ser atômicos
    with _RESERVA_LOCK:
        # Verificar se horário ainda está livre
        if not verificar_disponibilidade(data_str, hora_str):
            return {
                "sucesso": False,
                "chave": None,
                "expira_em": None,
                "mensagem": "Horário já está ocupado"
            }
        
        try:
            wb, ws = _open_ws()
            _ensure_headers(ws, HEADERS_AG)
            
            chave = make_key(data_str, hora_str, chat_id)
            reservado_em = datetime.now()
            reservado_ate = reservado_em + timedelta(minutes=duracao_reserva_min)
            
            # Criar linha com status "Reservado"
            _make_row(
                ws=ws,
                chave=chave,
                data_str=data_str,
                hora_str=hora_str,
                chat_id=chat_id,
                cliente_id=cliente_id,
                cliente_nome=cliente_nome,
                nasc=None,
                cpf=None,
                servico_id=servico_id,
                servico_duracao=servico_duracao,
                status="Reservado",
                reservado_em=reservado_em.isoformat(),
                reservado_ate=reservado_ate.isoformat(),
                valor_pago=None
            )
            
            wb.save(FILE_PATH)
            
            return {
                "sucesso": True,
                "chave": chave,
                "expira_em": reservado_ate.strftime("%d/%m/%Y %H:%M:%S"),
                "mensagem": f"Slot reservado até {reservado_ate.strftime('%H:%M')}"
            }
        
        except Exception as e:
            return {
                "sucesso": False,
                "chave": None,
                "expira_em": None,
                "mensagem": f"Erro ao reservar: {str(e)}"
            }


def confirmar_reserva(chave: str) -> bool:
    """
    Confirma uma reserva temporária, mudando status para "Confirmado".
    
    Args:
        chave: Chave do agendamento
    
    Returns:
        True se confirmado com sucesso
    """
    return atualizar_status_por_chave(chave, "Confirmado")


def cancelar_reserva(chave: str) -> bool:
    """
    Cancela uma reserva (libera o slot).
    
    Args:
        chave: Chave do agendamento
    
    Returns:
        True se cancelado com sucesso
    """
    return atualizar_status_por_chave(chave, "Cancelado")


def liberar_slots_expirados() -> int:
    """
    Busca reservas expiradas e libera os slots.
    
    Returns:
        Número de slots liberados
    """
    wb, ws = _open_ws()
    hm = _get_header_map(ws)
    
    c_status = hm.get("Status")
    c_reservado_ate = hm.get("ReservadoAte")
    
    if not c_status or not c_reservado_ate:
        return 0
    
    agora = datetime.now()
    liberados = 0
    
    for r in range(2, ws.max_row + 1):
        status = str(ws.cell(row=r, column=c_status).value or "").strip()
        
        # Apenas processar reservas pendentes
        if status != "Reservado":
            continue
        
        reservado_ate_str = str(ws.cell(row=r, column=c_reservado_ate).value or "").strip()
        
        if not reservado_ate_str:
            continue
        
        try:
            # Parse ISO format
            reservado_ate = datetime.fromisoformat(reservado_ate_str)
            
            # Se expirou
            if agora > reservado_ate:
                ws.cell(row=r, column=c_status, value="Expirado")
                liberados += 1
        
        except Exception:
            continue
    
    if liberados > 0:
        wb.save(FILE_PATH)
    
    return liberados


def verificar_reserva_ativa(data_str: str, hora_str: str, chat_id: str) -> bool:
    """
    Verifica se existe uma reserva ativa (não expirada) para este slot e cliente.
    
    Returns:
        True se o cliente tem reserva ativa neste slot
    """
    chave = make_key(data_str, hora_str, chat_id)
    
    wb, ws = _open_ws()
    hm = _get_header_map(ws)
    
    c_chave = hm.get("Chave")
    c_status = hm.get("Status")
    c_reservado_ate = hm.get("ReservadoAte")
    
    if not all([c_chave, c_status, c_reservado_ate]):
        return False
    
    for r in range(2, ws.max_row + 1):
        row_chave = str(ws.cell(row=r, column=c_chave).value or "").strip()
        
        if row_chave != chave:
            continue
        
        status = str(ws.cell(row=r, column=c_status).value or "").strip()
        
        if status != "Reservado":
            return False
        
        reservado_ate_str = str(ws.cell(row=r, column=c_reservado_ate).value or "").strip()
        
        try:
            reservado_ate = datetime.fromisoformat(reservado_ate_str)
            agora = datetime.now()
            
            # Reserva ainda válida
            return agora <= reservado_ate
        
        except Exception:
            return False
    
    return False


def obter_agendamentos_do_dia(data_str: str) -> List[Dict[str, Any]]:
    """
    Retorna todos os agendamentos de um dia específico.
    
    Args:
        data_str: Data no formato DD/MM/YYYY
    
    Returns:
        Lista de dicts com dados dos agendamentos
    """
    rows = _read_rows()
    return [r for r in rows if r.get("Data") == data_str]


def atualizar_pagamento_id(chave: str, payment_id: str, payment_status: str = "pending") -> bool:
    """
    Atualiza o PagamentoID e PagamentoStatus de um agendamento.
    
    Args:
        chave: Chave do agendamento
        payment_id: ID do pagamento no Mercado Pago
        payment_status: Status do pagamento (pending, approved, rejected)
    
    Returns:
        True se atualizado com sucesso
    """
    try:
        wb, ws = _open_ws()
        hm = _get_header_map(ws)
        
        c_chave = hm.get("Chave")
        c_pag_id = hm.get("PagamentoID")
        c_pag_status = hm.get("PagamentoStatus")
        
        if not all([c_chave, c_pag_id, c_pag_status]):
            return False
        
        for r in range(2, ws.max_row + 1):
            row_chave = str(ws.cell(row=r, column=c_chave).value or "").strip()
            
            if row_chave == chave:
                ws.cell(row=r, column=c_pag_id, value=str(payment_id))
                ws.cell(row=r, column=c_pag_status, value=payment_status)
                wb.save(FILE_PATH)
                return True
        
        return False
    
    except Exception as e:
        logger.error(f"Erro ao atualizar PagamentoID: {e}")
        return False
