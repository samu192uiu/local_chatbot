# services/excel_services.py
import os
from datetime import datetime, timedelta
from typing import List, Optional
from openpyxl import Workbook, load_workbook

# caminho do arquivo ao lado deste módulo (ou usar AGENDAMENTOS_XLSX se setado)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PLAN_PATH = os.getenv("AGENDAMENTOS_XLSX", os.path.join(BASE_DIR, "agendamentos.xlsx"))
SHEET = "Agendamentos"

COLS = [
    "Data", "Hora", "ClienteNome", "DataNascimento", "CPF",
    "ChatID", "Status", "ValorPago", "CriadoEm"
]

def _ensure_dirs():
    d = os.path.dirname(PLAN_PATH)
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)

def _ensure_sheet():
    _ensure_dirs()
    if not os.path.exists(PLAN_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET
        ws.append(COLS)
        wb.save(PLAN_PATH)
        return
    wb = load_workbook(PLAN_PATH)
    if SHEET not in wb.sheetnames:
        ws = wb.create_sheet(SHEET)
        ws.append(COLS)
        wb.save(PLAN_PATH)
    else:
        ws = wb[SHEET]
        if ws.max_row == 0:
            ws.append(COLS)
            wb.save(PLAN_PATH)
        else:
            headers = [c.value for c in ws[1]]
            if headers != COLS:
                ws.delete_rows(1, ws.max_row)
                ws.append(COLS)
                wb.save(PLAN_PATH)

def _read_rows():
    _ensure_sheet()
    wb = load_workbook(PLAN_PATH)
    ws = wb[SHEET]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(COLS, r)))
    wb.close()
    return rows

def _save_row(row: dict):
    _ensure_sheet()
    wb = load_workbook(PLAN_PATH)
    ws = wb[SHEET]
    values = [row.get(c) for c in COLS]
    ws.append(values)
    wb.save(PLAN_PATH)
    wb.close()

def _update_row(match_fn, update_fn) -> bool:
    _ensure_sheet()
    wb = load_workbook(PLAN_PATH)
    ws = wb[SHEET]
    updated = False
    for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        row_dict = {COLS[i]: r[i].value for i in range(len(COLS))}
        if match_fn(row_dict):
            update_fn(r)
            updated = True
            break
    if updated:
        wb.save(PLAN_PATH)
    wb.close()
    return updated

def _norm_hhmm(hhmm: str) -> Optional[str]:
    """Normaliza '8:00' -> '08:00' e valida."""
    try:
        dt = datetime.strptime(hhmm.strip(), "%H:%M")
        return dt.strftime("%H:%M")
    except Exception:
        return None

# ---------------------------
# API pública
# ---------------------------
def verificar_disponibilidade(data_str: str, hora_str: str) -> bool:
    """
    True se não houver entrada para Data/Hora com status != 'Cancelado'.
    """
    hhmm = _norm_hhmm(hora_str)
    if not hhmm:
        return False
    rows = _read_rows()
    for r in rows:
        if r["Data"] == data_str and (r["Hora"] or "") == hhmm:
            if (r.get("Status") or "").lower() != "cancelado":
                return False
    return True

def adicionar_agendamento(
    data_str: str,
    hora_str: str,
    chat_id: str,
    status: str = "Pendente Pagamento",
    cliente_nome: Optional[str] = None,
    data_nasc: Optional[str] = None,
    cpf: Optional[str] = None,
    valor_pago: Optional[float] = None,
):
    """
    Adiciona uma linha na planilha. Retorna uma chave única (data_hora_chat).
    """
    hhmm = _norm_hhmm(hora_str)
    if not hhmm:
        raise ValueError("Hora inválida (use HH:MM).")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = {
        "Data": data_str,
        "Hora": hhmm,
        "ClienteNome": cliente_nome,
        "DataNascimento": data_nasc,
        "CPF": cpf,
        "ChatID": chat_id,
        "Status": status,
        "ValorPago": valor_pago,
        "CriadoEm": now,
    }
    _save_row(row)
    return make_key(data_str, hhmm, chat_id)

def make_key(data_str: str, hora_str: str, chat_id: str) -> str:
    hhmm = _norm_hhmm(hora_str) or hora_str
    return f"{data_str}_{hhmm}_{chat_id}"

def atualizar_status_por_chave(*args) -> bool:
    """
    Compatível com duas assinaturas:
    (chave, novo_status)
    (data_str, hora_str, chat_id, novo_status)
    """
    if len(args) == 2:
        chave, novo_status = args
        def match_fn(r): return make_key(r["Data"], r["Hora"], r["ChatID"]) == chave
    elif len(args) == 4:
        data_str, hora_str, chat_id, novo_status = args
        hhmm = _norm_hhmm(hora_str)
        def match_fn(r): return r["Data"] == data_str and (r["Hora"] or "") == hhmm and r["ChatID"] == chat_id
    else:
        raise TypeError("Use (chave, status) ou (data, hora, chat_id, status).")

    def update_fn(cells):
        idx = COLS.index("Status")
        cells[idx].value = novo_status

    return _update_row(match_fn, update_fn)

def listar_horarios_disponiveis(
    data_str: str,
    inicio: str = "08:00",
    fim: str = "18:00",
    passo_min: int = 30,
    allowed_slots: Optional[List[str]] = None,
) -> List[str]:
    """
    Se allowed_slots for passado, usa exatamente aqueles horários (normalizados).
    Caso contrário, gera a grade [inicio..fim] de passo_min.
    Remove horários já ocupados (status != Cancelado).
    """
    rows = _read_rows()
    ocupados = set(
        (r["Hora"] or "")
        for r in rows
        if r["Data"] == data_str and (r.get("Status") or "").lower() != "cancelado"
    )

    if allowed_slots:
        pool = []
        for h in allowed_slots:
            nh = _norm_hhmm(h)
            if nh and nh not in pool:
                pool.append(nh)
        # mantém ordem fornecida
        return [h for h in pool if h not in ocupados]

    # fallback: grade automática
    base_dt = datetime.strptime(data_str, "%d/%m/%Y")
    t_ini = datetime.strptime(inicio, "%H:%M")
    t_fim = datetime.strptime(fim, "%H:%M")
    cursor = datetime.combine(base_dt.date(), t_ini.time())
    end = datetime.combine(base_dt.date(), t_fim.time())

    out = []
    while cursor <= end:
        hstr = cursor.strftime("%H:%M")
        if hstr not in ocupados:
            out.append(hstr)
        cursor += timedelta(minutes=passo_min)
    return out
