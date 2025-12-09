# src/zapwaha/web/webhooks.py
from __future__ import annotations
import os
import json
import logging
import requests
from collections import deque
from typing import Iterable, Optional, Any, Dict

from flask import Blueprint, request, jsonify
from zapwaha.flows.agendamento import route_message

# Client opcional (não dependemos dele para enviar)
try:
    from services.waha import Waha  # opcional
except Exception:
    Waha = None

# Sistema de lembretes
try:
    from services import reminders
    _REMINDERS_ENABLED = True
except Exception as e:
    reminders = None
    _REMINDERS_ENABLED = False
    logging.warning(f"Sistema de lembretes não disponível: {e}")

# ====== Config (ENV) ======
WAHA_API_URL = os.getenv("WAHA_API_URL", "http://waha:3000").rstrip("/")
WAHA_SESSION = (os.getenv("WAHA_SESSION", "default") or "default").strip()
WAHA_API_KEY = os.getenv("WAHA_API_KEY")  # se setada, enviamos em X-Api-Key
ADMIN_TOKEN = os.getenv("ADMIN_TOKEN")  # para endpoints de debug

web_bp = Blueprint("web", __name__)
logger = logging.getLogger("ZapWaha")
if not logger.handlers:
    logger.setLevel(logging.INFO)

# Deduplicação: agora guardamos **somente event_ids reais**
_RECENT_EVENT_IDS: deque[str] = deque(maxlen=2000)

# Flag para controlar inicialização única dos lembretes
_LEMBRETES_INICIADOS = False

# --------------------------------------------------------------------------------------
# Utilitários
# --------------------------------------------------------------------------------------
def _asdict(x: Any) -> Dict[str, Any]:
    return x if isinstance(x, dict) else {}

def _log_payload(data: Any, note: str = ""):
    try:
        txt = json.dumps(data, ensure_ascii=False)[:1200]
        prefix = f"[WEBHOOK]{' ' + note if note else ''}"
        logger.debug(f"{prefix} payload: {txt}")
    except Exception:
        pass

def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple, set)):
        return ""
    try:
        return str(value)
    except Exception:
        return ""

# --------------------------------------------------------------------------------------
# Client opcional
# --------------------------------------------------------------------------------------
def build_waha():
    if not Waha:
        return None
    try:
        return Waha(api_url=WAHA_API_URL)
    except TypeError:
        pass
    try:
        return Waha(WAHA_API_URL)
    except TypeError:
        pass
    try:
        obj = Waha()
        for attr in ("api_url", "base_url", "url", "_Waha__api_url"):
            try:
                setattr(obj, attr, WAHA_API_URL)
            except Exception:
                pass
        return obj
    except Exception:
        return None

waha_service = build_waha()

# --------------------------------------------------------------------------------------
# Envio de mensagens (HTTP direto + client opcional)
# --------------------------------------------------------------------------------------
def _send_http_waha(chat_id: str, message: str) -> bool:
    base = WAHA_API_URL
    session = WAHA_SESSION

    headers_base = {}
    if WAHA_API_KEY:
        headers_base["X-Api-Key"] = WAHA_API_KEY

    url = f"{base}/api/sendText"
    variations = [
        (url, {"session": session, "chatId": chat_id, "text": message}, None),
        (url, {"session": session, "chatId": chat_id, "message": message}, None),

        (url, {"chatId": chat_id, "text": message}, {"x-session": session}),
        (url, {"chatId": chat_id, "message": message}, {"x-session": session}),
        (url, {"chatId": chat_id, "text": message}, {"X-Session": session}),
        (url, {"chatId": chat_id, "message": message}, {"X-Session": session}),

        (f"{url}?session={session}", {"chatId": chat_id, "text": message}, None),
        (f"{url}?session={session}", {"chatId": chat_id, "message": message}, None),

        (url, {"session": session, "to": chat_id, "text": message}, None),
        (f"{url}?session={session}", {"to": chat_id, "text": message}, None),

        (f"{base}/api/sessions/{session}/messages/text", {"chatId": chat_id, "text": message}, None),
        (f"{base}/api/messages/send", {"session": session, "chatId": chat_id, "text": message}, None),
    ]

    for u, b, extra_headers in variations:
        try:
            headers = {**headers_base, **(extra_headers or {})}
            r = requests.post(u, json=b, headers=headers, timeout=10)
            if 200 <= r.status_code < 300:
                logger.info(f"[WAHA HTTP] OK via {u}")
                return True
            logger.warning(f"[WAHA HTTP] {u} -> {r.status_code} {r.text[:200]}")
        except Exception as e:
            logger.warning(f"[WAHA HTTP] erro em {u}: {e}")
    return False

def _send(chat_id: str, message: str):
    if _send_http_waha(chat_id, message):
        return

    if waha_service:
        try:
            if hasattr(waha_service, "send_message"):
                try:
                    waha_service.send_message(chat_id=chat_id, message=message)
                    return
                except TypeError:
                    waha_service.send_message(chat_id, message)
                    return
            if hasattr(waha_service, "send_text"):
                try:
                    waha_service.send_text(chat_id=chat_id, text=message)
                    return
                except TypeError:
                    waha_service.send_text(chat_id, message)
                    return
        except Exception as e:
            logger.error(f"Falha ao enviar via WAHA (client): {e}")

    print(f"[SEND-FALLBACK to {chat_id}] {message}")


# --------------------------------------------------------------------------------------
# Inicialização do sistema de lembretes
# --------------------------------------------------------------------------------------
def _inicializar_lembretes_se_necessario():
    """Inicializa o sistema de lembretes uma única vez."""
    global _LEMBRETES_INICIADOS
    
    if _LEMBRETES_INICIADOS:
        return
    
    if not _REMINDERS_ENABLED or not reminders:
        return
    
    try:
        # Inicializar com a função de envio
        reminders.inicializar_lembretes(_send)
        _LEMBRETES_INICIADOS = True
        logger.info("✅ Sistema de lembretes automáticos iniciado")
    except Exception as e:
        logger.error(f"❌ Erro ao inicializar sistema de lembretes: {e}")


# --------------------------------------------------------------------------------------
# typing (best-effort)
# --------------------------------------------------------------------------------------
def _typing_http(chat_id: str, on: bool):
    try:
        endpoint = "/api/startTyping" if on else "/api/stopTyping"
        headers = {}
        if WAHA_API_KEY:
            headers["X-Api-Key"] = WAHA_API_KEY
        headers["X-Session"] = WAHA_SESSION
        r = requests.post(
            f"{WAHA_API_URL}{endpoint}",
            json={"session": WAHA_SESSION, "chatId": chat_id},
            headers=headers,
            timeout=4,
        )
        if not (200 <= r.status_code < 300):
            logger.debug(f"[WAHA HTTP] typing {endpoint} -> {r.status_code} {r.text[:120]}")
    except Exception:
        pass

def _typing(chat_id: str, on: bool):
    _typing_http(chat_id, on)
    if not waha_service:
        return
    try:
        if on and hasattr(waha_service, "start_typing"):
            try:
                waha_service.start_typing(chat_id=chat_id)
            except TypeError:
                waha_service.start_typing(chat_id)
        if (not on) and hasattr(waha_service, "stop_typing"):
            try:
                waha_service.stop_typing(chat_id=chat_id)
            except TypeError:
                waha_service.stop_typing(chat_id)
    except Exception:
        pass

# --------------------------------------------------------------------------------------
# Parsing helpers
# --------------------------------------------------------------------------------------
def _normalize_chat(chat_id: Optional[str]) -> Optional[str]:
    if not chat_id:
        return None
    chat_id = str(chat_id).strip()
    if "@s.whatsapp.net" in chat_id:
        chat_id = chat_id.replace("@s.whatsapp.net", "@c.us")
    if chat_id.isdigit():
        chat_id = f"{chat_id}@c.us"
    return chat_id

def _extract_event_id(payload: Any) -> Optional[str]:
    p = _asdict(payload)

    # Preferir IDs de mensagem (mais específicos)
    msg = p.get("message") or _asdict(p.get("data")).get("message")
    if isinstance(msg, dict):
        mid = msg.get("id") or _asdict(msg.get("key")).get("id") or msg.get("messageId")
        if mid:
            return f"msg:{mid}"

    # IDs de event, se existirem
    ev = p.get("event")
    if isinstance(ev, dict):
        eid = ev.get("id") or ev.get("eventId")
        if eid:
            return f"evt:{eid}"
    for parent in ("payload", "data"):
        inner = _asdict(p.get(parent)).get("event")
        if isinstance(inner, dict):
            eid = inner.get("id") or inner.get("eventId")
            if eid:
                return f"evt:{eid}"

    # NÃO usar id top-level genérico
    return None

def _extract_chat_id(payload: Any) -> Optional[str]:
    p = _asdict(payload)

    cid = (
        _asdict(p.get("payload")).get("from")
        or _asdict(p.get("data")).get("from")
        or p.get("from")
        or p.get("sender")
        or p.get("chatId")
        or p.get("phone")
        or _asdict(p.get("data")).get("chatId")
        or _asdict(p.get("data")).get("phone")
        or _asdict(p.get("data")).get("to")
    )

    if not cid:
        msg = p.get("message") or _asdict(p.get("data")).get("message")
        if isinstance(msg, dict):
            cid = (
                msg.get("from")
                or msg.get("chatId")
                or _asdict(msg.get("key")).get("remoteJid")
                or msg.get("remoteJid")
                or msg.get("to")
            )

    return _normalize_chat(cid)

def _extract_text(payload: Any) -> str:
    p = _asdict(payload)

    candidates = [
        _asdict(p.get("payload")).get("body"),
        _asdict(p.get("payload")).get("text"),
        _asdict(p.get("payload")).get("caption"),
        _asdict(_asdict(p.get("data")).get("msg")).get("text"),
        _asdict(p.get("data")).get("text"),
        _asdict(p.get("data")).get("body"),
        _asdict(p.get("data")).get("caption"),
    ]

    msg = p.get("message") or _asdict(p.get("data")).get("message")
    if isinstance(msg, dict):
        candidates.extend([
            _asdict(msg).get("text"),
            _asdict(msg).get("caption"),
            _asdict(msg).get("conversation"),
        ])
        text_obj = _asdict(msg.get("text"))
        if text_obj:
            candidates.append(text_obj.get("text"))

    candidates.extend([
        p.get("body"),
        p.get("caption"),
        p.get("text"),
    ])

    for v in candidates:
        s = _as_text(v)
        if s.strip():
            return s
    return ""

def _event_type(payload: Any) -> str:
    p = _asdict(payload)
    ev = p.get("event")
    if isinstance(ev, dict):
        return _as_text(ev.get("type") or ev.get("eventType")).lower()
    if isinstance(ev, str):
        return ev.lower()
    for parent in ("payload", "data"):
        inner = _asdict(p.get(parent)).get("event")
        if isinstance(inner, dict):
            return _as_text(inner.get("type") or inner.get("eventType")).lower()
        if isinstance(inner, str):
            return inner.lower()
    return _as_text(p.get("type")).lower()

def _session_matches(payload: Any) -> bool:
    p = _asdict(payload)
    sess = (
        p.get("session")
        or _asdict(p.get("payload")).get("session")
        or _asdict(p.get("data")).get("session")
    )
    return not sess or str(sess) == WAHA_SESSION

# --------------------------------------------------------------------------------------
# Dedup: **somente por event_id real** (se não tiver, não deduplica)
# --------------------------------------------------------------------------------------
def _seen_event(event_id: Optional[str]) -> bool:
    if not event_id:
        return False
    if event_id in _RECENT_EVENT_IDS:
        return True
    _RECENT_EVENT_IDS.append(event_id)
    return False

# --------------------------------------------------------------------------------------
# Rotas principais
# --------------------------------------------------------------------------------------
@web_bp.get("/healthz")
def healthz():
    return jsonify({"ok": True, "session": WAHA_SESSION}), 200

@web_bp.post("/chatbot/webhook/")
def chatbot_webhook():
    try:
        # Inicializar lembretes (só na primeira chamada)
        _inicializar_lembretes_se_necessario()
        
        payload = request.get_json(silent=True)
        if isinstance(payload, str):
            try:
                payload = json.loads(payload)
            except Exception:
                payload = {"_raw": payload}
        payload = payload or {}
        _log_payload(payload)

        processed = 0
        ignored = 0
        deduped = 0

        for unit in _iter_units(payload):
            try:
                if not isinstance(unit, dict):
                    ignored += 1
                    continue

                if not _session_matches(unit):
                    ignored += 1
                    continue

                chat_id = _extract_chat_id(unit)
                text = _extract_text(unit).strip()
                if not chat_id or not text:
                    et = _event_type(unit)
                    if et in {"message_ack", "message-status", "ack", "status", "delivery", "read"}:
                        ignored += 1
                        continue
                    ignored += 1
                    continue

                evt_id = _extract_event_id(unit)
                if _seen_event(evt_id):
                    deduped += 1
                    continue

                logger.debug(f"[WEBHOOK] in chat_id={chat_id} evt_id={evt_id} text={text[:60]!r}")

                _typing(chat_id, True)
                route_message(_send, chat_id, text)
                _typing(chat_id, False)
                processed += 1

            except Exception as inner_e:
                logger.exception(f"[WEBHOOK] erro ao processar unidade: {inner_e}")
                try:
                    cid = _extract_chat_id(unit) if isinstance(unit, dict) else None
                    if cid:
                        _typing(cid, False)
                        _send(cid, "Desculpe, ocorreu um erro interno. Digite 'menu' para recomeçar.")
                except Exception:
                    pass

        logger.info(f"[WEBHOOK] processed={processed} ignored={ignored} deduped={deduped}")
        return jsonify({"status": "success", "processed": processed, "ignored": ignored, "deduped": deduped}), 200

    except Exception as e:
        logger.exception(f"Erro no webhook: {e}")
        try:
            chat_id_on_error = _extract_chat_id(request.get_json(silent=True) or {})
            if chat_id_on_error:
                _typing(chat_id_on_error, False)
                _send(chat_id_on_error, "Desculpe, ocorreu um erro interno. Digite 'menu' para recomeçar.")
        except Exception:
            pass
        return jsonify({"status": "error", "message": "Erro interno (handled)."}), 200

@web_bp.post("/test/incoming")
def test_incoming():
    payload = request.get_json(silent=True) or {}
    chat_id = str(payload.get("user_id", "5511912345678@c.us"))
    text = _as_text(payload.get("text", ""))

    chat_id = _normalize_chat(chat_id)

    _typing(chat_id, True)
    route_message(_send, chat_id, text.strip())
    _typing(chat_id, False)

    return jsonify({"ok": True, "echo": {"chat_id": chat_id, "text": text}}), 200

# --------------------------------------------------------------------------------------
# Explode lista de mensagens do envelope
# --------------------------------------------------------------------------------------
def _iter_units(payload: Any) -> Iterable[Any]:
    # LISTA de eventos no corpo
    if isinstance(payload, list):
        for item in payload:
            yield item
        return

    # Qualquer coisa não-dict (str etc.)
    if not isinstance(payload, dict):
        yield payload
        return

    pp = _asdict(payload)
    payload_block = _asdict(pp.get("payload"))
    data_block = _asdict(pp.get("data"))

    msgs = None
    if isinstance(payload_block.get("messages"), list):
        msgs = payload_block["messages"]
    elif isinstance(data_block.get("messages"), list):
        msgs = data_block["messages"]
    elif isinstance(pp.get("messages"), list):
        msgs = pp["messages"]

    if msgs is None:
        yield payload
    else:
        for m in msgs:
            if isinstance(m, dict):
                unit = {"message": m}
                for k in ("payload", "data", "session", "event", "id", "type"):
                    if k in pp:
                        unit[k] = pp[k]
                yield unit
            else:
                yield m

# ======================================================================================
# ==========================  ENDPOINTS DE DEBUG / ADMIN  ==============================
# ======================================================================================
# Para consultar e manipular clientes sem abrir a planilha manualmente.

# Dependências do módulo de clientes
from services import clientes_services as CS
from openpyxl import load_workbook

def _require_admin():
    if not ADMIN_TOKEN:
        return jsonify({"error": "ADMIN_TOKEN não configurado"}), 501
    token = request.headers.get("X-Admin-Token") or request.args.get("token")
    if token != ADMIN_TOKEN:
        return jsonify({"error": "forbidden"}), 403
    return None  # ok

@web_bp.get("/debug/clients/lookup")
def debug_clients_lookup():
    """?chatId=...  ou  ?cpf=..."""
    guard = _require_admin()
    if guard:
        return guard

    chat = request.args.get("chatId")
    cpf = request.args.get("cpf")

    if chat:
        chat_norm = _normalize_chat(chat)
        rec = CS.get_by_chat_id(chat_norm)
        return jsonify({"ok": True, "by": "chatId", "chatId_norm": chat_norm, "record": rec}), 200

    if cpf:
        rec = CS.get_by_cpf(cpf)
        return jsonify({"ok": True, "by": "cpf", "record": rec}), 200

    return jsonify({"ok": False, "error": "informe chatId ou cpf"}), 400

@web_bp.get("/debug/clients/list")
def debug_clients_list():
    """Lista registros da planilha de clientes (paginado)."""
    guard = _require_admin()
    if guard:
        return guard

    try:
        limit = int(request.args.get("limit", 50))
        offset = int(request.args.get("offset", 0))
    except Exception:
        return jsonify({"ok": False, "error": "limit/offset inválidos"}), 400

    CS.init_planilha()
    path = CS.FILE_PATH
    try:
        wb = load_workbook(path)
        ws = wb.active
        rows = []
        start = max(2 + offset, 2)
        end = min(ws.max_row, start + limit - 1)
        for r in range(start, end + 1):
            item = {}
            for idx, h in enumerate(CS.HEADERS, 1):
                item[h] = ws.cell(row=r, column=idx).value
            rows.append(item)
        return jsonify({"ok": True, "file": path, "count": len(rows), "rows": rows}), 200
    except Exception as e:
        return jsonify({"ok": False, "file": path, "error": str(e)}), 500

@web_bp.post("/debug/clients/upsert")
def debug_clients_upsert():
    """Body JSON: {CPF, Nome, Nascimento, Telefone, Email, ChatId}"""
    guard = _require_admin()
    if guard:
        return guard

    data = request.get_json(silent=True) or {}
    if "ChatId" in data:
        data["ChatId"] = _normalize_chat(data.get("ChatId"))

    try:
        rec = CS.create_or_update_client(data)
        return jsonify({"ok": True, "record": rec}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@web_bp.delete("/debug/clients/by-cpf")
def debug_clients_delete_by_cpf():
    """Remove cliente por CPF exato (linha inteira)."""
    guard = _require_admin()
    if guard:
        return guard

    cpf = request.args.get("cpf")
    if not cpf:
        return jsonify({"ok": False, "error": "informe ?cpf="}), 400

    CS.init_planilha()
    path = CS.FILE_PATH
    try:
        wb = load_workbook(path)
        ws = wb.active
        # Localiza coluna do CPF
        try:
            cpf_col = CS.HEADERS.index("CPF") + 1
        except ValueError:
            return jsonify({"ok": False, "error": "Cabeçalho sem coluna CPF"}), 500

        row_to_delete = None
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=cpf_col).value or "").strip() == str(cpf).strip():
                row_to_delete = r
                break
        if not row_to_delete:
            return jsonify({"ok": False, "error": "CPF não encontrado"}), 404

        ws.delete_rows(row_to_delete, 1)
        wb.save(path)
        return jsonify({"ok": True, "deleted_row": row_to_delete}), 200
    except Exception as e:
        return jsonify({"ok": False, "file": path, "error": str(e)}), 500
