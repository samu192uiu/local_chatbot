# src/zapwaha/web/debug.py
from __future__ import annotations
import os
import logging
from typing import Any, Dict, Iterable, List, Optional

from flask import Blueprint, request, jsonify, current_app

logger = logging.getLogger("ZapWaha")
debug_bp = Blueprint("debug_clients", __name__)

# ===== Auth =====
def _expected_token() -> Optional[str]:
    # 1) ENV; 2) app.config como fallback
    return os.getenv("ADMIN_TOKEN") or current_app.config.get("ADMIN_TOKEN")

def _auth_ok(req) -> bool:
    token = None
    auth = req.headers.get("Authorization", "")
    if auth.lower().startswith("bearer "):
        token = auth.split(" ", 1)[1].strip()
    if not token:
        token = req.headers.get("X-Admin-Token") or req.args.get("token")
    exp = _expected_token()
    return bool(exp and token == exp)

@debug_bp.before_request
def _check_admin_token():
    # libera preflight
    if request.method == "OPTIONS":
        return
    if not _auth_ok(request):
        return jsonify({"error": "unauthorized"}), 401


# ===== Data sources =====
# 1) clientes_services se existir
try:
    from services import clientes_services as _clientes
except Exception:
    _clientes = None

def _iter_clients_from_service() -> List[Dict[str, Any]]:
    """
    Tenta diferentes APIs comuns no módulo clientes_services.
    Normaliza campos chaves: Nome, CPF, Nascimento, Telefone, Email, ChatId, UltimoLogin.
    """
    rows: Iterable[Dict[str, Any]] = []
    try:
        if not _clientes:
            return []
        if hasattr(_clientes, "list_all"):
            rows = _clientes.list_all()  # esperado: list[dict]
        elif hasattr(_clientes, "get_all_clients"):
            rows = _clientes.get_all_clients()
        elif hasattr(_clientes, "read_rows"):
            rows = _clientes.read_rows()
        elif hasattr(_clientes, "get_all"):
            rows = _clientes.get_all()
        else:
            return []
    except Exception as e:
        logger.warning(f"[DEBUG] erro lendo clientes do service: {e}")
        return []

    out: List[Dict[str, Any]] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        # normalização de chaves
        nome   = r.get("Nome") or r.get("nome")
        cpf    = r.get("CPF") or r.get("cpf")
        nasc   = r.get("Nascimento") or r.get("nascimento") or r.get("DataNascimento")
        fone   = r.get("Telefone") or r.get("telefone") or r.get("Phone")
        email  = r.get("Email") or r.get("email")
        chat   = r.get("ChatId") or r.get("chat_id") or r.get("ChatID")
        ultimo = r.get("UltimoLogin") or r.get("ultimo_login") or r.get("LastLogin")
        out.append({
            "Nome": nome, "CPF": cpf, "Nascimento": nasc,
            "Telefone": fone, "Email": email, "ChatId": chat, "UltimoLogin": ultimo
        })
    return out


# 2) Fallback direto no XLSX (se openpyxl estiver disponível)
def _read_xlsx_clients() -> List[Dict[str, Any]]:
    path = os.getenv("CLIENTES_XLSX") or "/app/data/clientes.xlsx"
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        # Cabeçalho na 1ª linha
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}

        def _get(row, name):
            i = idx.get(name)
            if i is None:
                # tenta variações de nome
                aliases = {
                    "Nome": ["nome"],
                    "CPF": ["cpf"],
                    "Nascimento": ["nascimento", "DataNascimento"],
                    "Telefone": ["telefone", "Phone"],
                    "Email": ["email"],
                    "ChatId": ["chat_id", "ChatID"],
                    "UltimoLogin": ["ultimo_login", "LastLogin"],
                }
                for alt in aliases.get(name, []):
                    i = idx.get(alt)
                    if i is not None:
                        break
            if i is None:
                return None
            v = row[i].value
            return str(v).strip() if v is not None else None

        out: List[Dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            rec = {
                "Nome": _get(row, "Nome"),
                "CPF": _get(row, "CPF"),
                "Nascimento": _get(row, "Nascimento"),
                "Telefone": _get(row, "Telefone"),
                "Email": _get(row, "Email"),
                "ChatId": _get(row, "ChatId"),
                "UltimoLogin": _get(row, "UltimoLogin"),
            }
            # ignora linhas totalmente vazias
            if any(rec.values()):
                out.append(rec)
        return out
    except Exception as e:
        logger.warning(f"[DEBUG] erro lendo {path}: {e}")
        return []


def _load_clients() -> List[Dict[str, Any]]:
    rows = _iter_clients_from_service()
    if rows:
        return rows
    # fallback XLSX
    return _read_xlsx_clients()


def _mask_cpf(cpf: Optional[str]) -> Optional[str]:
    if not cpf:
        return cpf
    s = "".join(ch for ch in cpf if ch.isdigit())
    if len(s) == 11:
        return f"{s[:3]}.{s[3:6]}.{s[6:9]}-{s[9:]}"
    return cpf


# ===== Endpoints =====
@debug_bp.get("/logins")
def list_logins():
    """
    Retorna "logins ativos" = clientes com ChatId preenchido (ou UltimoLogin recente, se existir).
    """
    rows = _load_clients()
    ativos = []
    for r in rows:
        chat = r.get("ChatId")
        ultimo = r.get("UltimoLogin")
        if chat or ultimo:
            ativos.append({
                "ChatId": chat,
                "CPF": _mask_cpf(r.get("CPF")),
                "Nome": r.get("Nome"),
                "UltimoLogin": ultimo
            })
    return jsonify({"count": len(ativos), "items": ativos}), 200


@debug_bp.get("/all")
def list_all_clients():
    """
    Lista todos os clientes conhecidos (normalizados).
    """
    rows = _load_clients()
    # máscara leve de CPF
    for r in rows:
        r["CPF"] = _mask_cpf(r.get("CPF"))
    return jsonify({"count": len(rows), "items": rows}), 200


@debug_bp.get("/cpf/<cpf>")
def get_by_cpf(cpf: str):
    """
    Busca um cliente por CPF (aceita com ou sem pontuação).
    """
    needle = "".join(ch for ch in (cpf or "") if ch.isdigit())
    rows = _load_clients()
    for r in rows:
        doc = "".join(ch for ch in (r.get("CPF") or "") if ch.isdigit())
        if doc and doc == needle:
            rec = dict(r)
            rec["CPF"] = _mask_cpf(rec.get("CPF"))
            return jsonify(rec), 200
    return jsonify({"error": "not found"}), 404
